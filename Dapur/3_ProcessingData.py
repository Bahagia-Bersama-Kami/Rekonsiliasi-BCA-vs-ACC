import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import timedelta
import itertools
import re

def clean_currency(x):
    if pd.isna(x): return 0.0
    if isinstance(x, (int, float)): return float(x)
    x = str(x).strip()
    if x == '': return 0.0
    if '.' in x and ',' in x:
        x = x.replace('.', '').replace(',', '.')
    elif ',' in x:
        x = x.replace(',', '.')
    elif '.' in x:
        if x.count('.') > 1 or len(x.split('.')[-1]) == 3:
            x = x.replace('.', '')
    try:
        return float(x)
    except:
        return 0.0

def get_matching_indices(indices, values, target):
    if not values: return None
    if abs(sum(values) - target) < 1.5:
        return list(indices)
    
    items = sorted(zip(indices, values), key=lambda x: x[1], reverse=True)
    iters = [0]
    
    def backtrack(start, current_sum, path):
        iters[0] += 1
        if iters[0] > 300000: return None
        
        if abs(current_sum - target) < 1.5:
            return path
        
        if current_sum > target + 1.5:
            return None
        
        for i in range(start, len(items)):
            if len(path) >= 25: continue
            res = backtrack(i + 1, current_sum + items[i][1], path + [items[i][0]])
            if res: return res
        return None

    return backtrack(0, 0, [])

acc_df = pd.read_excel('Acc_temp.xlsx')
bca_df = pd.read_excel('Bca_temp.xlsx')

acc_df['Tanggal'] = pd.to_datetime(acc_df['Tanggal'], dayfirst=True, format='mixed', errors='coerce')
bca_df['Tanggal Transaksi'] = pd.to_datetime(bca_df['Tanggal Transaksi'], dayfirst=True, format='mixed', errors='coerce')

acc_df['Penambahan'] = acc_df['Penambahan'].apply(clean_currency)
acc_df['Pengurangan'] = acc_df['Pengurangan'].apply(clean_currency)
bca_df['Kredit'] = bca_df['Kredit'].apply(clean_currency)
bca_df['Debet'] = bca_df['Debet'].apply(clean_currency)

acc_df['Matched'] = False
bca_df['Matched'] = False

acc_df['Key_Search'] = acc_df['Keterangan'].astype(str).apply(lambda x: x.split('for ')[-1].strip() if 'for ' in x else x)

matched_data = []

for max_days in [0, 5]:
    for b_idx, b_row in bca_df[~bca_df['Matched']].iterrows():
        b_date = b_row['Tanggal Transaksi']
        if pd.isna(b_date): continue
        val = b_row['Kredit'] if b_row['Kredit'] > 0 else -b_row['Debet']
        if val == 0: continue
        col_acc = 'Penambahan' if val > 0 else 'Pengurangan'
        target_val = abs(val)
        potential = acc_df[(~acc_df['Matched']) & (abs(acc_df[col_acc] - target_val) < 1) & ((acc_df['Tanggal'] - b_date).dt.days.abs() <= max_days)]
        
        if not potential.empty:
            potential = potential.iloc[(potential['Tanggal'] - b_date).dt.days.abs().argsort()]
            a_idx = None
            b_ket = str(b_row['Keterangan']).upper()
            b_words = set(re.findall(r'[A-Z0-9]{3,}', b_ket))
            
            for p_idx, p_row in potential.iterrows():
                a_ket = str(p_row['Keterangan']).upper()
                a_words = set(re.findall(r'[A-Z0-9]{3,}', a_ket))
                if b_words & a_words:
                    a_idx = p_idx
                    break
                    
            if a_idx is None:
                a_idx = potential.index[0]
                
            acc_df.at[a_idx, 'Matched'] = True
            bca_df.at[b_idx, 'Matched'] = True
            matched_data.append({'Tanggal Bank': b_date, 'Tanggal Admin': acc_df.at[a_idx, 'Tanggal'], 'Keterangan Bank': b_row['Keterangan'], 'Keterangan Admin': acc_df.at[a_idx, 'Keterangan'], 'No. Sumber': acc_df.at[a_idx, 'No. Sumber'], 'Nominal': target_val, 'Tipe': '1:1 Match'})

    for b_idx, b_row in bca_df[~bca_df['Matched']].iterrows():
        b_date = b_row['Tanggal Transaksi']
        val = b_row['Kredit'] if b_row['Kredit'] > 0 else -b_row['Debet']
        col_acc = 'Penambahan' if val > 0 else 'Pengurangan'
        target_val = abs(val)
        potential_acc = acc_df[(~acc_df['Matched']) & (acc_df[col_acc] > 0) & ((acc_df['Tanggal'] - b_date).dt.days.abs() <= max_days)]
        if potential_acc.empty: continue
        groups = potential_acc.groupby('Key_Search')
        for name, group in groups:
            matched_idx = get_matching_indices(group.index.tolist(), group[col_acc].tolist(), target_val)
            if matched_idx:
                matched_rows = acc_df.loc[matched_idx]
                acc_df.loc[matched_idx, 'Matched'] = True
                bca_df.at[b_idx, 'Matched'] = True
                matched_data.append({'Tanggal Bank': b_date, 'Tanggal Admin': matched_rows['Tanggal'].iloc[0], 'Keterangan Bank': b_row['Keterangan'], 'Keterangan Admin': ", ".join(matched_rows['Keterangan'].astype(str).tolist()), 'No. Sumber': ", ".join(matched_rows['No. Sumber'].astype(str).tolist()), 'Nominal': target_val, 'Tipe': 'Group Match (Acc to BCA)'})
                break

    for a_idx, a_row in acc_df[~acc_df['Matched']].iterrows():
        a_date = a_row['Tanggal']
        col_bca = 'Kredit' if a_row['Penambahan'] > 0 else 'Debet'
        target_val = a_row['Penambahan'] if a_row['Penambahan'] > 0 else a_row['Pengurangan']
        if target_val == 0: continue
        potential_bca = bca_df[(~bca_df['Matched']) & (bca_df[col_bca] > 0) & ((bca_df['Tanggal Transaksi'] - a_date).dt.days.abs() <= max_days)]
        if potential_bca.empty: continue
        matched_idx_bca = get_matching_indices(potential_bca.index.tolist(), potential_bca[col_bca].tolist(), target_val)
        if matched_idx_bca:
            matched_rows_bca = bca_df.loc[matched_idx_bca]
            bca_df.loc[matched_idx_bca, 'Matched'] = True
            acc_df.at[a_idx, 'Matched'] = True
            matched_data.append({'Tanggal Bank': matched_rows_bca['Tanggal Transaksi'].iloc[0], 'Tanggal Admin': a_date, 'Keterangan Bank': ", ".join(matched_rows_bca['Keterangan'].astype(str).tolist()), 'Keterangan Admin': a_row['Keterangan'], 'No. Sumber': a_row['No. Sumber'], 'Nominal': target_val, 'Tipe': 'Group Match (BCA to Acc)'})

    unmatched_bca_last = bca_df[~bca_df['Matched']].index
    for b_idx in unmatched_bca_last:
        b_row = bca_df.loc[b_idx]
        val = b_row['Kredit'] if b_row['Kredit'] > 0 else -b_row['Debet']
        col_acc = 'Penambahan' if val > 0 else 'Pengurangan'
        target_val = abs(val)
        potential_acc = acc_df[(~acc_df['Matched']) & (acc_df[col_acc] > 0) & ((acc_df['Tanggal'] - b_row['Tanggal Transaksi']).dt.days.abs() <= max_days)]
        if not potential_acc.empty:
            matched_idx = get_matching_indices(potential_acc.index.tolist(), potential_acc[col_acc].tolist(), target_val)
            if matched_idx:
                matched_rows = acc_df.loc[matched_idx]
                acc_df.loc[matched_idx, 'Matched'] = True
                bca_df.at[b_idx, 'Matched'] = True
                matched_data.append({'Tanggal Bank': b_row['Tanggal Transaksi'], 'Tanggal Admin': matched_rows['Tanggal'].iloc[0], 'Keterangan Bank': b_row['Keterangan'], 'Keterangan Admin': ", ".join(matched_rows['Keterangan'].astype(str).tolist()), 'No. Sumber': ", ".join(matched_rows['No. Sumber'].astype(str).tolist()), 'Nominal': target_val, 'Tipe': 'General Group Match'})

summary_data = pd.DataFrame({'A. Keterangan': ['Total Acc Penambahan', 'Total Bca Kredit', 'Selisih Penambahan', 'Total Acc Pengurangan', 'Total Bca Debet', 'Selisih Pengurangan'], 'Nominal': [acc_df['Penambahan'].sum(), bca_df['Kredit'].sum(), abs(acc_df['Penambahan'].sum() - bca_df['Kredit'].sum()), acc_df['Pengurangan'].sum(), bca_df['Debet'].sum(), abs(acc_df['Pengurangan'].sum() - bca_df['Debet'].sum())]})
unmatched_acc = acc_df[~acc_df['Matched']].copy()
unmatched_bca = bca_df[~bca_df['Matched']].copy()
unmatched_acc['Keterangan Rekonsiliasi'] = 'Hanya ada di Accurate'
unmatched_bca['Keterangan Rekonsiliasi'] = 'Hanya ada di BCA'
unmatched_bca.rename(columns={'Tanggal Transaksi': 'Tanggal', 'Kredit': 'Penambahan', 'Debet': 'Pengurangan'}, inplace=True)
unmatched_all = pd.concat([unmatched_acc[['Tanggal', 'No. Sumber', 'Keterangan', 'Penambahan', 'Pengurangan', 'Keterangan Rekonsiliasi']], unmatched_bca[['Tanggal', 'Keterangan', 'Penambahan', 'Pengurangan', 'Keterangan Rekonsiliasi']]], ignore_index=True)
unmatched_all.sort_values(by=['Keterangan Rekonsiliasi', 'Tanggal'], ascending=[True, True], inplace=True)
df_matched = pd.DataFrame(matched_data)
if not df_matched.empty: df_matched.sort_values(by='Tanggal Bank', inplace=True)

with pd.ExcelWriter('Hasil_Rekonsiliasi.xlsx', engine='openpyxl') as writer:
    summary_data.to_excel(writer, sheet_name='Analisis', index=False, startrow=0)
    start_unmatched = len(summary_data) + 3
    pd.DataFrame([['B. Hasil Rekonsiliasi (Tidak Sesuai / Selisih)']]).to_excel(writer, sheet_name='Analisis', index=False, header=False, startrow=start_unmatched - 1)
    if not unmatched_all.empty: unmatched_all.to_excel(writer, sheet_name='Analisis', index=False, startrow=start_unmatched)
    start_matched = start_unmatched + len(unmatched_all) + 4
    pd.DataFrame([['C. Data Yang Sesuai']]).to_excel(writer, sheet_name='Analisis', index=False, header=False, startrow=start_matched - 1)
    if not df_matched.empty: df_matched.to_excel(writer, sheet_name='Analisis', index=False, startrow=start_matched)
    acc_df.drop(columns=['Matched', 'Key_Search'], errors='ignore').to_excel(writer, sheet_name='Acc_temp Asli', index=False)
    bca_df.drop(columns=['Matched'], errors='ignore').to_excel(writer, sheet_name='Bca_temp Asli', index=False)

wb = load_workbook('Hasil_Rekonsiliasi.xlsx')

ws = wb['Analisis']
bold_font = Font(bold=True)
for cell in ws[1]: cell.font = bold_font
ws.cell(row=1, column=1).font = bold_font
ws.cell(row=start_unmatched, column=1).font = bold_font
ws.cell(row=start_matched, column=1).font = bold_font
fill_summary, fill_unmatched, fill_matched = PatternFill(start_color='EAF2F8', end_color='EAF2F8', fill_type='solid'), PatternFill(start_color='FDEDEC', end_color='FDEDEC', fill_type='solid'), PatternFill(start_color='EAFAF1', end_color='EAFAF1', fill_type='solid')
for row in ws.iter_rows(min_row=2, max_row=len(summary_data)+1, min_col=1, max_col=2):
    for cell in row:
        cell.fill = fill_summary
        if cell.column == 2: cell.number_format = '#,##0'
if not unmatched_all.empty:
    h_row = start_unmatched + 1
    for cell in ws[h_row]: cell.font = bold_font
    for row in ws.iter_rows(min_row=h_row+1, max_row=h_row+len(unmatched_all), min_col=1, max_col=6):
        for cell in row:
            cell.fill = fill_unmatched
            if cell.column == 1: cell.number_format = 'DD/MM/YYYY'
            if cell.column in [4, 5]: cell.number_format = '#,##0'
if not df_matched.empty:
    h_row = start_matched + 1
    for cell in ws[h_row]: cell.font = bold_font
    for row in ws.iter_rows(min_row=h_row+1, max_row=h_row+len(df_matched), min_col=1, max_col=7):
        for cell in row:
            cell.fill = fill_matched
            if cell.column in [1, 2]: cell.number_format = 'DD/MM/YYYY'
            if cell.column == 6: cell.number_format = '#,##0'
kolom_lebar = {'A': 15, 'B': 20, 'C': 45, 'D': 45, 'E': 25, 'F': 20, 'G': 25}
for col, width in kolom_lebar.items(): ws.column_dimensions[col].width = width

if 'Acc_temp Asli' in wb.sheetnames:
    ws_acc = wb['Acc_temp Asli']
    kolom_lebar_acc = {'A': 18, 'B': 13, 'C': 80, 'D': 15, 'E': 15, 'F': 15}
    for col, width in kolom_lebar_acc.items(): ws_acc.column_dimensions[col].width = width

if 'Bca_temp Asli' in wb.sheetnames:
    ws_bca = wb['Bca_temp Asli']
    kolom_lebar_bca = {'A': 18, 'B': 70, 'C': 9, 'D': 15, 'E': 15, 'F': 15}
    for col, width in kolom_lebar_bca.items(): ws_bca.column_dimensions[col].width = width

wb.save('Hasil_Rekonsiliasi.xlsx')